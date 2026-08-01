#!/usr/bin/env python3
"""
prep-irish-crt.py

Stage 1 of the two-stage rule: read Ireland's published Common Reporting Tables and write
a small tidy CSV that the manuscript consumes with base R read.csv.

Source
------
IRL-CRT-2026-V1.0, Ireland's 2026 submission to the UNFCCC, 35 annual workbooks 1990-2024.
Downloaded 2026-08-01 from
  https://www.epa.ie/publications/monitoring--assessment/climate-change/air-emissions/IRL-CRT-2026-V1.0.zip
(HTTP 200, 40,816,340 bytes). Kept at 02.inputs/IRL/CRT/. The UNFCCC mirror sits behind a
WAF that returns a JavaScript challenge to curl; use the EPA copy.

What this computes, and why it is not simply read off one table
--------------------------------------------------------------
Ireland splits the balance for a hectare of organic soil across two CRT tables:

  Table 4.D      net carbon stock change in the SOIL, t C/ha, on-site. Sign convention is
                 the CRT's: positive = gain (removal), negative = loss (emission).
  Table 4(II)    emissions and removals from drainage and rewetting: an off-site CO2 term
                 (for Ireland this is the DOC flux), N2O-N, and CH4, all per hectare.

Neither table alone is the balance. Adding them is the whole point, and it is the arithmetic
the guidance never requires anyone to perform.

IMPORTANT, and the reason this script exists rather than a quoted figure
-----------------------------------------------------------------------
A prior review derived these numbers using the NID's *stated* emission factor of 1.7
t CO2-C/ha/yr for industrial peat extraction. The CRT's own *implied* factor for 2024 is
3.587 t C/ha, more than twice that, and the difference reverses the sign of the GWP-100
result. Implied factors are what the inventory actually reports, being emissions divided by
area, so they are what this script uses. Anything quoted from the earlier derivation is
superseded.

Metrics
-------
Ireland reports at AR5 GWP-100 (CH4 = 28, N2O = 265). AR6 values are computed alongside for
comparability with the rest of the manuscript, which uses AR6 Table 7.15 throughout.
"""

import csv
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: python3 -m pip install --user openpyxl")

IN_DIR = "../02.inputs/IRL/CRT"
OUT = "../02.inputs/derived/irish-crt-peat-balance.csv"

C_TO_CO2 = 44 / 12
N_TO_N2O = 44 / 28

METRICS = {
    "AR5_GWP100": {"ch4": 28.0, "n2o": 265.0},   # what Ireland reports at
    "AR6_GWP100": {"ch4": 27.0, "n2o": 273.0},   # AR6 Table 7.15, biogenic CH4
    "AR6_GWP20":  {"ch4": 79.7, "n2o": 273.0},
}

# Rows are located by their label text in column B rather than by row number, because row
# positions are not guaranteed stable across 35 workbooks.
LABELS_4D = {
    "industrial_drained": "Industrial Peat Extraction",
    "industrial_rewet":   "Rewetted Industrial Peat Extraction",
    "domestic_drained":   "Domestic Peat Extraction",
    "near_natural":       "Near Natural Wetlands",
}
LABELS_4II = {
    "industrial_drained": "Industrial Peat Extraction",
    "industrial_rewet":   "Rewetted Industrial Peat Extraction",
    "domestic_drained":   "Domestic Peat Extraction",
}


def num(v):
    """CRT cells carry notation keys (NO, NA, IE, NE) as strings. Those are not zero."""
    if isinstance(v, (int, float)):
        return float(v)
    return None


def find_rows(ws, label, maxrow=120):
    """All rows whose column-B label matches exactly, in order."""
    hits = []
    for r in range(1, maxrow + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip() == label:
            hits.append(r)
    return hits


def read_year(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Table4.D" not in wb.sheetnames or "Table4(II)" not in wb.sheetnames:
        return None
    d, ii = wb["Table4.D"], wb["Table4(II)"]

    out = {}

    # Table 4.D: column D = total area (kha), column L = organic-soil net C stock change.
    for key, label in LABELS_4D.items():
        rows = find_rows(d, label)
        if not rows:
            continue
        r = rows[0]
        out[key] = {
            "area_kha_4D": num(d.cell(row=r, column=4).value),
            "soilC_tC_ha": num(d.cell(row=r, column=12).value),
        }

    # Table 4(II): D = area (kha), E = CO2 kg/ha, F = N2O-N kg/ha, G = CH4 kg/ha.
    # The same label appears under both 4(II).D.1.a (peat extraction remaining) and the
    # flooded-land block, where every value is the notation key "NO". Take the first row
    # that carries a real area.
    for key, label in LABELS_4II.items():
        for r in find_rows(ii, label):
            a = num(ii.cell(row=r, column=4).value)
            if a is None:
                continue
            out.setdefault(key, {}).update({
                "area_kha_4II": a,
                "co2_kg_ha":    num(ii.cell(row=r, column=5).value),
                "n2o_n_kg_ha":  num(ii.cell(row=r, column=6).value),
                "ch4_kg_ha":    num(ii.cell(row=r, column=7).value),
            })
            break

    # Near-natural wetlands sit under 4(II).D.1.c, whose per-hectare row is labelled
    # "Total organic soils"; locate it by the block header above it.
    for r in range(1, 120):
        v = ii.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip().startswith("4(II).D.1.c"):
            for rr in range(r + 1, r + 4):
                if str(ii.cell(row=rr, column=2).value).strip() == "Total organic soils":
                    out.setdefault("near_natural", {}).update({
                        "area_kha_4II": num(ii.cell(row=rr, column=4).value),
                        "co2_kg_ha":    num(ii.cell(row=rr, column=5).value),
                        "n2o_n_kg_ha":  num(ii.cell(row=rr, column=6).value),
                        "ch4_kg_ha":    num(ii.cell(row=rr, column=7).value),
                    })
                    break
            break
    return out


def balance(rec, metric):
    """Total t CO2e per hectare per year. Positive = net source."""
    g = METRICS[metric]
    if rec.get("soilC_tC_ha") is None:
        return None
    # Table 4.D sign is positive-for-gain; flip so positive means emission.
    onsite = -rec["soilC_tC_ha"] * C_TO_CO2
    offsite = (rec.get("co2_kg_ha") or 0.0) / 1000.0
    n2o = ((rec.get("n2o_n_kg_ha") or 0.0) / 1000.0) * N_TO_N2O * g["n2o"]
    ch4 = ((rec.get("ch4_kg_ha") or 0.0) / 1000.0) * g["ch4"]
    return onsite + offsite + n2o + ch4


def main():
    files = sorted(
        f for f in os.listdir(IN_DIR)
        if re.match(r"IRL-CRT-.*-(\d{4})\.xlsx$", f) and not f.startswith("~")
    )
    if not files:
        sys.exit(f"No CRT workbooks in {IN_DIR}. See the download URL in this file's docstring.")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    rows = []
    for f in files:
        year = int(re.search(r"(\d{4})\.xlsx$", f).group(1))
        rec = read_year(os.path.join(IN_DIR, f))
        if not rec:
            continue
        for key, r in rec.items():
            row = {
                "year": year, "category": key,
                "area_kha_4D": r.get("area_kha_4D"),
                "area_kha_4II": r.get("area_kha_4II"),
                "soilC_tC_ha": r.get("soilC_tC_ha"),
                "co2_kg_ha": r.get("co2_kg_ha"),
                "n2o_n_kg_ha": r.get("n2o_n_kg_ha"),
                "ch4_kg_ha": r.get("ch4_kg_ha"),
            }
            for m in METRICS:
                row[f"total_{m}"] = balance(r, m)
            rows.append(row)

    cols = ["year", "category", "area_kha_4D", "area_kha_4II", "soilC_tC_ha",
            "co2_kg_ha", "n2o_n_kg_ha", "ch4_kg_ha"] + [f"total_{m}" for m in METRICS]
    with open(OUT, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({c: ("" if r.get(c) is None else r.get(c)) for c in cols})

    print(f"Wrote {OUT}: {len(rows)} rows, {len(files)} years "
          f"({min(r['year'] for r in rows)}-{max(r['year'] for r in rows)})")

    latest = max(r["year"] for r in rows)
    cur = {r["category"]: r for r in rows if r["year"] == latest}
    print(f"\n--- {latest}, t CO2e/ha/yr, positive = net source ---")
    for m in METRICS:
        dr = cur.get("industrial_drained", {}).get(f"total_{m}")
        rw = cur.get("industrial_rewet", {}).get(f"total_{m}")
        if dr is None or rw is None:
            continue
        print(f"  {m:11s} drained {dr:7.3f}  rewetted {rw:7.3f}  "
              f"benefit {dr - rw:+7.3f}")
    nn = cur.get("near_natural", {})
    if nn.get("total_AR5_GWP100") is not None:
        print(f"\n  Near-natural wetland, {latest}: on-site soil C "
              f"{nn['soilC_tC_ha']:+.3f} t C/ha (a removal), but the full balance is "
              f"{nn['total_AR5_GWP100']:+.3f} t CO2e/ha at AR5 GWP-100 over "
              f"{nn['area_kha_4D']:.0f} kha.")

    # Area inconsistency between the two tables, worth reporting rather than smoothing over.
    ind = cur.get("industrial_drained", {})
    if ind.get("area_kha_4D") and ind.get("area_kha_4II"):
        gap = ind["area_kha_4D"] - ind["area_kha_4II"]
        print(f"\n  NOTE: industrial peat extraction area is {ind['area_kha_4D']:.2f} kha in "
              f"Table 4.D but {ind['area_kha_4II']:.2f} kha in Table 4(II); "
              f"{gap:.2f} kha carries an on-site CO2 term but no CH4 or N2O.")


if __name__ == "__main__":
    main()
